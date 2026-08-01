"""Smoke tests for the FastAPI web app.

These tests guard against regressions in the TemplateResponse API usage.
Starlette 1.0 changed TemplateResponse to require `request` as the first
positional argument.  Passing a string template-name first (old API) causes
``TypeError: unhashable type: 'dict'`` / ``AttributeError: 'dict' object has
no attribute 'split'`` at runtime on a clean server installation.
"""
from __future__ import annotations

import os
import re
import sqlite3
from datetime import date
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from app.services.stock_xlsx import generate_stock_xlsx_bytes
from app.services.invoice_xlsx import generate_invoice_xlsx_bytes
from app.services.return_xlsx import generate_return_xlsx_bytes
from app.services.receive_xlsx import generate_receive_xlsx_bytes
from app.utils.money import calc_line_total


@pytest.fixture(scope="module")
def client(tmp_path_factory: pytest.TempPathFactory) -> TestClient:
    data_dir = tmp_path_factory.mktemp("data")
    db_path = data_dir / "stock.db"
    os.environ["DB_PATH"] = str(db_path)
    os.environ["BACKUP_DIR"] = str(data_dir / "backups")
    os.environ["INVOICES_DIR"] = str(data_dir / "invoices")

    # Import *after* setting DB_PATH so the module picks up the correct path.
    from app.web.main import app  # noqa: PLC0415

    with TestClient(app, raise_server_exceptions=True) as c:
        yield c


def test_index_returns_200(client: TestClient) -> None:
    """GET / must return 200 (not 500) on a clean install."""
    response = client.get("/")
    assert response.status_code == 200, (
        f"GET / returned {response.status_code}; "
        "possible TemplateResponse API mismatch with current Starlette version"
    )
    assert "text/html" in response.headers["content-type"]
    assert "Hasapcy" in response.text


def test_products_page_returns_200(client: TestClient) -> None:
    """GET /products must render without error."""
    response = client.get("/products")
    assert response.status_code == 200


def test_index_nav_groups_and_reports_ru(client: TestClient) -> None:
    """GET / in RU must expose compact nav groups and the renamed Help item."""
    response = client.get("/", headers={"cookie": "ui_lang=ru"})
    assert response.status_code == 200
    assert 'class="navbar navbar-expand-lg navbar-dark bg-dark app-top-nav"' in response.text
    assert "Отчёты" in response.text
    assert "Документы" in response.text
    assert "Справочники" in response.text
    assert "Помощь" in response.text
    assert 'href="/reports"' in response.text
    assert 'href="/invoices"' in response.text
    assert 'href="/history"' in response.text
    assert 'href="/brands"' in response.text
    assert 'href="/clients"' in response.text
    assert 'href="/suppliers"' in response.text
    assert 'href="/help">Помощь<' in response.text
    assert 'href="/help">Справочник<' not in response.text


def test_reports_page_returns_basic_analytics_ru(client: TestClient) -> None:
    """GET /reports in RU must render the working basic analytics page."""
    response = client.get("/reports", headers={"cookie": "ui_lang=ru"})
    assert response.status_code == 200
    assert "📊 Отчёты" in response.text
    assert "Базовая аналитика по продажам, возвратам и складу" in response.text
    assert "Выручка за период" in response.text
    assert "Топ товаров по продажам" in response.text
    assert "Низкие остатки" in response.text
    assert "id=\"salesChart\"" in response.text


def test_reports_page_custom_period_uses_sales_returns_and_stock_data(client: TestClient) -> None:
    from app.db.sqlite import DB_PATH

    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute(
            "INSERT OR IGNORE INTO warehouses (code, title) VALUES (?, ?)",
            ("RPT_WH", "Report Warehouse"),
        )
        conn.execute(
            "INSERT INTO clients (name, phone, note, created_at, archived) VALUES (?, '', '', datetime('now','localtime'), 0)",
            ("Reports Test Client",),
        )
        client_id = int(conn.execute("SELECT id FROM clients WHERE name = ?", ("Reports Test Client",)).fetchone()[0])

        conn.execute(
            "INSERT INTO products (brand, model, name, wh_price, barcode, note, archived) VALUES (?, ?, ?, ?, '', '', 0)",
            ("RPT_BRAND_A", "RPT_MODEL_A", "Report Product A", 30.0),
        )
        conn.execute(
            "INSERT INTO products (brand, model, name, wh_price, barcode, note, archived) VALUES (?, ?, ?, ?, '', '', 0)",
            ("RPT_BRAND_B", "RPT_MODEL_B", "Report Product B", 60.0),
        )
        product_a_id = int(
            conn.execute(
                "SELECT id FROM products WHERE brand = ? AND model = ?",
                ("RPT_BRAND_A", "RPT_MODEL_A"),
            ).fetchone()[0]
        )
        product_b_id = int(
            conn.execute(
                "SELECT id FROM products WHERE brand = ? AND model = ?",
                ("RPT_BRAND_B", "RPT_MODEL_B"),
            ).fetchone()[0]
        )

        conn.execute(
            "INSERT INTO stock (warehouse_code, product_id, qty) VALUES (?, ?, ?)",
            ("RPT_WH", product_a_id, 1.0),
        )
        conn.execute(
            "INSERT INTO stock (warehouse_code, product_id, qty) VALUES (?, ?, ?)",
            ("RPT_WH", product_b_id, 10.0),
        )

        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, created_at, status) VALUES (?, ?, ?, 'CLOSED')",
            (client_id, "RPT_WH", "2026-01-05 10:00:00"),
        )
        cart_in_period = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        conn.execute(
            "INSERT INTO invoices (cart_id, number, created_at, currency, total) VALUES (?, ?, ?, 'USD', ?)",
            (cart_in_period, 900001, "2026-01-05 10:00:00", 100.0),
        )
        conn.execute(
            "INSERT INTO cart_items (cart_id, product_id, free_line, free_name, qty, price_mode, unit_price, total) VALUES (?, ?, 0, '', ?, 'custom', ?, ?)",
            (cart_in_period, product_a_id, 2.0, 50.0, 100.0),
        )

        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, created_at, status) VALUES (?, ?, ?, 'CLOSED')",
            (client_id, "RPT_WH", "2026-02-10 10:00:00"),
        )
        cart_out_period = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        conn.execute(
            "INSERT INTO invoices (cart_id, number, created_at, currency, total) VALUES (?, ?, ?, 'USD', ?)",
            (cart_out_period, 900002, "2026-02-10 10:00:00", 200.0),
        )
        conn.execute(
            "INSERT INTO cart_items (cart_id, product_id, free_line, free_name, qty, price_mode, unit_price, total) VALUES (?, ?, 0, '', ?, 'custom', ?, ?)",
            (cart_out_period, product_b_id, 1.0, 200.0, 200.0),
        )

        conn.execute(
            """
            INSERT INTO return_invoices (number, client_id, warehouse_code, status, created_at, currency, total, note)
            VALUES (?, ?, ?, 'DONE', ?, 'USD', ?, '')
            """,
            (910001, client_id, "RPT_WH", "2026-01-06 15:00:00", 20.0),
        )
        return_invoice_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        conn.execute(
            """
            INSERT INTO return_items (invoice_id, product_id, qty, unit_price, total)
            VALUES (?, ?, ?, ?, ?)
            """,
            (return_invoice_id, product_a_id, 1.0, 20.0, 20.0),
        )
        conn.commit()

    response = client.get(
        "/reports?date_from=2026-01-01&date_to=2026-01-31",
        headers={"cookie": "ui_lang=ru"},
    )
    assert response.status_code == 200
    assert "Выручка за период" in response.text
    assert "100.00 $" in response.text
    assert "Возвраты за период" in response.text
    assert "20.00 $" in response.text
    assert "Чистая выручка" in response.text
    assert "80.00 $" in response.text
    assert "Количество продаж / чеков" in response.text
    assert "RPT_BRAND_A" in response.text
    assert "RPT_MODEL_A" in response.text
    assert "RPT_WH" in response.text
    assert "2026-01-05" in response.text


def test_reports_page_preset_period_overrides_stale_date_params(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    from app.web import main as web_main

    class FixedDate(date):
        @classmethod
        def today(cls) -> date:
            return cls(2026, 3, 15)

    monkeypatch.setattr(web_main, "date", FixedDate)

    response = client.get(
        "/reports?period=today&date_from=2026-01-01&date_to=2026-01-31",
        headers={"cookie": "ui_lang=ru"},
    )
    assert response.status_code == 200
    assert 'name="date_from" value="2026-03-15"' in response.text
    assert 'name="date_to" value="2026-03-15"' in response.text
    assert '<option value="custom"' in response.text


def test_reports_page_contains_period_date_sync_script(client: TestClient) -> None:
    response = client.get("/reports", headers={"cookie": "ui_lang=ru"})
    assert response.status_code == 200
    assert 'id="reportsPeriod"' in response.text
    assert "computeRangeFromPreset" in response.text
    assert "periodSelect.addEventListener('change'" in response.text
    assert "periodSelect.value = 'custom'" in response.text


def test_reports_all_time_period_preset_present(client: TestClient) -> None:
    """The all_time preset option must appear in all supported languages."""
    for lang, expected_label in [("ru", "За всё время"), ("en", "All time"), ("tm", "Tutuş döwür")]:
        response = client.get("/reports", headers={"cookie": f"ui_lang={lang}"})
        assert response.status_code == 200
        assert 'value="all_time"' in response.text, f"all_time option missing for lang={lang}"
        assert expected_label in response.text, f"Label missing for lang={lang}"


def test_reports_all_time_period_overrides_stale_date_params(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    """period=all_time must override stale date_from/date_to params."""
    from app.web import main as web_main

    class FixedDate(date):
        @classmethod
        def today(cls) -> date:
            return cls(2026, 5, 12)

    monkeypatch.setattr(web_main, "date", FixedDate)

    response = client.get(
        "/reports?period=all_time&date_from=2099-01-01&date_to=2099-12-31",
        headers={"cookie": "ui_lang=ru"},
    )
    assert response.status_code == 200
    # date_to must be today
    assert 'name="date_to" value="2026-05-12"' in response.text
    # date_from must NOT be the stale 2099 date
    assert 'name="date_from" value="2099-01-01"' not in response.text


def test_reports_all_time_js_handler_present(client: TestClient) -> None:
    """The JS computeRangeFromPreset must contain an all_time branch."""
    response = client.get("/reports", headers={"cookie": "ui_lang=ru"})
    assert response.status_code == 200
    assert "all_time" in response.text
    assert "earliest" in response.text


def test_reports_warehouse_filter_applies_to_stock_kpis_and_low_stock(client: TestClient) -> None:
    from app.db.sqlite import DB_PATH

    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.executemany(
            "INSERT OR IGNORE INTO warehouses (code, title) VALUES (?, ?)",
            [("RWF_A", "Reports Warehouse A"), ("RWF_B", "Reports Warehouse B")],
        )
        conn.executemany(
            "INSERT INTO products (brand, model, name, wh_price, barcode, note, archived) VALUES (?, ?, ?, ?, '', '', 0)",
            [
                ("RWF_BRAND", "RWF_HIGH_A", "Reports Filter High A", 10.0),
                ("RWF_BRAND", "RWF_HIGH_B", "Reports Filter High B", 11.0),
                ("RWF_BRAND", "RWF_LOW_A", "Reports Filter Low A", 12.0),
                ("RWF_BRAND", "RWF_LOW_B", "Reports Filter Low B", 13.0),
            ],
        )
        rows = conn.execute(
            "SELECT id, model FROM products WHERE model IN (?, ?, ?, ?)",
            ("RWF_HIGH_A", "RWF_HIGH_B", "RWF_LOW_A", "RWF_LOW_B"),
        ).fetchall()
        product_ids = {model: pid for pid, model in rows}
        conn.executemany(
            "INSERT INTO stock (warehouse_code, product_id, qty) VALUES (?, ?, ?)",
            [
                ("RWF_A", product_ids["RWF_HIGH_A"], 5.0),
                ("RWF_B", product_ids["RWF_HIGH_B"], 7.0),
                ("RWF_A", product_ids["RWF_LOW_A"], 1.0),
                ("RWF_B", product_ids["RWF_LOW_B"], 1.0),
            ],
        )
        conn.commit()

    response = client.get(
        "/reports?date_from=2026-01-01&date_to=2026-01-31&warehouses=RWF_A",
        headers={"cookie": "ui_lang=ru"},
    )
    assert response.status_code == 200
    assert 'option value="RWF_A" selected' in response.text
    assert 'option value="RWF_B" selected' not in response.text
    assert 'option value="all" selected' not in response.text
    assert "RWF_LOW_A" in response.text
    assert "RWF_LOW_B" not in response.text

    stock_qty_match = re.search(
        r"Товаров на складе, шт\.</div><div class=\"h5 mb-0\">([0-9.]+)</div>",
        response.text,
    )
    assert stock_qty_match
    assert float(stock_qty_match.group(1)) == 6.0

    positions_match = re.search(
        r"Позиций с остатком</div><div class=\"h5 mb-0\">([0-9]+)</div>",
        response.text,
    )
    assert positions_match
    assert int(positions_match.group(1)) == 2


def test_reports_low_stock_sorted_by_warehouse_brand_model(client: TestClient) -> None:
    from app.db.sqlite import DB_PATH

    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.executemany(
            "INSERT OR IGNORE INTO warehouses (code, title) VALUES (?, ?)",
            [("SRT_A", "Sort Warehouse A"), ("SRT_B", "Sort Warehouse B")],
        )
        conn.executemany(
            "INSERT INTO products (brand, model, name, wh_price, barcode, note, archived) VALUES (?, ?, ?, ?, '', '', 0)",
            [
                ("SRT_BRAND_A", "SRT_MODEL_1", "Sort Product 1", 10.0),
                ("SRT_BRAND_A", "SRT_MODEL_0", "Sort Product 0", 10.0),
                ("SRT_BRAND_B", "SRT_MODEL_2", "Sort Product 2", 10.0),
            ],
        )
        rows = conn.execute(
            "SELECT id, model FROM products WHERE model IN (?, ?, ?)",
            ("SRT_MODEL_1", "SRT_MODEL_0", "SRT_MODEL_2"),
        ).fetchall()
        product_ids = {model: pid for pid, model in rows}
        conn.executemany(
            "INSERT INTO stock (warehouse_code, product_id, qty) VALUES (?, ?, ?)",
            [
                ("SRT_B", product_ids["SRT_MODEL_2"], 1.0),
                ("SRT_A", product_ids["SRT_MODEL_1"], 1.0),
                ("SRT_A", product_ids["SRT_MODEL_0"], 1.0),
            ],
        )
        conn.commit()

    response = client.get(
        "/reports?date_from=2026-01-01&date_to=2026-01-31",
        headers={"cookie": "ui_lang=ru"},
    )
    assert response.status_code == 200
    idx_model_0 = response.text.find("SRT_MODEL_0")
    idx_model_1 = response.text.find("SRT_MODEL_1")
    idx_model_2 = response.text.find("SRT_MODEL_2")
    assert idx_model_0 != -1 and idx_model_1 != -1 and idx_model_2 != -1
    assert idx_model_0 < idx_model_1 < idx_model_2


def test_reports_low_stock_sorted_by_qty_ascending(client: TestClient) -> None:
    """Low-stock table must sort by qty ASC first so the most critical items appear first."""
    from app.db.sqlite import DB_PATH

    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.executemany(
            "INSERT OR IGNORE INTO warehouses (code, title) VALUES (?, ?)",
            [("QTY_WH", "Qty Sort Warehouse")],
        )
        conn.executemany(
            "INSERT INTO products (brand, model, name, wh_price, barcode, note, archived) VALUES (?, ?, ?, ?, '', '', 0)",
            [
                ("QTY_BRAND", "QTY_MODEL_TWO", "Qty Product 2", 10.0),
                ("QTY_BRAND", "QTY_MODEL_ZERO", "Qty Product 0", 10.0),
                ("QTY_BRAND", "QTY_MODEL_ONE", "Qty Product 1", 10.0),
            ],
        )
        rows = conn.execute(
            "SELECT id, model FROM products WHERE model IN (?, ?, ?)",
            ("QTY_MODEL_TWO", "QTY_MODEL_ZERO", "QTY_MODEL_ONE"),
        ).fetchall()
        product_ids = {model: pid for pid, model in rows}
        conn.executemany(
            "INSERT INTO stock (warehouse_code, product_id, qty) VALUES (?, ?, ?)",
            [
                ("QTY_WH", product_ids["QTY_MODEL_TWO"], 2.0),
                ("QTY_WH", product_ids["QTY_MODEL_ZERO"], 0.0),
                ("QTY_WH", product_ids["QTY_MODEL_ONE"], 1.0),
            ],
        )
        conn.commit()

    response = client.get(
        "/reports?date_from=2026-01-01&date_to=2026-01-31",
        headers={"cookie": "ui_lang=ru"},
    )
    assert response.status_code == 200
    idx_zero = response.text.find("QTY_MODEL_ZERO")
    idx_one = response.text.find("QTY_MODEL_ONE")
    idx_two = response.text.find("QTY_MODEL_TWO")
    assert idx_zero != -1 and idx_one != -1 and idx_two != -1
    # qty=0 must come before qty=1, which must come before qty=2
    assert idx_zero < idx_one < idx_two


@pytest.mark.parametrize(
    ("lang", "manual_entry_text"),
    [
        ("en", "or enter/scan barcode"),
        ("ru", "или введите/отсканируйте штрихкод"),
    ],
)
def test_catalog_page_supports_manual_barcode_scanner_entry(
    client: TestClient, lang: str, manual_entry_text: str
) -> None:
    """GET /catalog must expose Enter handling and scanner-friendly copy."""
    response = client.get("/catalog", headers={"cookie": f"ui_lang={lang}"})
    assert response.status_code == 200
    assert manual_entry_text in response.text
    assert "addEventListener('keydown'" in response.text
    assert "event.key === 'Enter'" in response.text
    assert "focusManualBarcodeSoon();" in response.text


def test_index_background_has_dynamic_cache_busting_version(
    client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from app.web import main as web_main

    static_dir = tmp_path / "static"
    static_dir.mkdir()
    bg_path = static_dir / "bg.jpg"
    bg_path.write_bytes(b"main-bg")
    os.utime(bg_path, (1_715_260_000, 1_715_260_000))
    monkeypatch.setattr(web_main, "STATIC_DIR", static_dir)

    response = client.get("/")
    assert response.status_code == 200
    assert f'/static/bg.jpg?v={bg_path.stat().st_mtime_ns}' in response.text


def test_price_page_background_has_dynamic_cache_busting_version(
    client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from app.web import main as web_main

    static_dir = tmp_path / "static"
    static_dir.mkdir()
    bg_path = static_dir / "price-bg.jpg"
    bg_path.write_bytes(b"price-bg")
    os.utime(bg_path, (1_715_260_100, 1_715_260_100))
    monkeypatch.setattr(web_main, "STATIC_DIR", static_dir)

    response = client.get("/price")
    assert response.status_code == 200
    assert f'/static/price-bg.jpg?v={bg_path.stat().st_mtime_ns}' in response.text


def test_admin_price_background_upload_overwrites_price_bg_file(
    client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from app.web import main as web_main

    static_dir = tmp_path / "static"
    static_dir.mkdir()
    monkeypatch.setattr(web_main, "STATIC_DIR", static_dir)

    response = client.post(
        "/admin/settings/price-background",
        files={"price_bg_file": ("new-price-bg.webp", b"new-image", "image/webp")},
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/admin/settings?saved=1"
    assert (static_dir / "price-bg.jpg").read_bytes() == b"new-image"


def test_unlock_page_returns_200(client: TestClient) -> None:
    """GET /unlock must render without error."""
    response = client.get("/unlock")
    assert response.status_code == 200
    assert "Hasapcy" in response.text


def test_help_page_includes_suppliers_section_ru(client: TestClient) -> None:
    """GET /help in RU must include Suppliers TOC item and section anchor."""
    response = client.get("/help", headers={"cookie": "ui_lang=ru"})
    assert response.status_code == 200
    assert 'href="#ru-suppliers"' in response.text
    assert 'id="ru-suppliers"' in response.text
    assert 'href="#ru-reports"' in response.text
    assert 'id="ru-reports"' in response.text
    assert "Структура верхнего меню" in response.text
    assert "Помощь ERP" in response.text


def test_help_page_includes_suppliers_section_en(client: TestClient) -> None:
    """GET /help in EN must include Suppliers TOC item and section anchor."""
    response = client.get("/help", headers={"cookie": "ui_lang=en"})
    assert response.status_code == 200
    assert 'href="#en-suppliers"' in response.text
    assert 'id="en-suppliers"' in response.text
    assert 'href="#en-reports"' in response.text
    assert 'id="en-reports"' in response.text
    assert "Top menu structure" in response.text


def test_stock_xlsx_returns_xlsx_bytes(client: TestClient) -> None:
    """GET /stock/xlsx must return an XLSX file (not 500) even with empty warehouse."""
    response = client.get("/stock/xlsx?warehouse=&q=&show_archived=0")
    assert response.status_code == 200, (
        f"GET /stock/xlsx returned {response.status_code}; expected 200"
    )
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # XLSX files start with PK (ZIP magic bytes)
    assert response.content[:2] == b"PK"


def test_generate_stock_xlsx_bytes_missing_warehouse_key() -> None:
    """generate_stock_xlsx_bytes must not crash when rows lack a 'warehouse' key."""
    rows = [
        {
            "brand": "Nike",
            "model": "Air Max",
            "name": "Sneaker",
            "warehouse_code": "TM_DEPO",
            "qty": 5,
        }
    ]
    result = generate_stock_xlsx_bytes(rows, "TM_DEPO")
    assert result[:2] == b"PK", "Expected XLSX (ZIP) bytes"


def test_generate_stock_xlsx_bytes_both_keys_missing() -> None:
    """generate_stock_xlsx_bytes must not crash even when both warehouse keys are missing."""
    rows = [
        {
            "brand": "Adidas",
            "model": "Stan Smith",
            "name": "Shoe",
            "qty": 3,
        }
    ]
    result = generate_stock_xlsx_bytes(rows, "All Warehouses")
    assert result[:2] == b"PK", "Expected XLSX (ZIP) bytes"


def test_generate_invoice_xlsx_bytes_includes_barcode() -> None:
    """generate_invoice_xlsx_bytes must include a Barcode column and not crash."""
    import openpyxl, io

    invoice = {
        "number": 1,
        "client": "Test Client",
        "created_at": "2024-01-01T10:00:00",
        "total": 200.0,
    }
    items = [
        {
            "brand": "Nike",
            "model": "Air Max",
            "name": "Sneaker",
            "barcode": "1234567890",
            "qty": 2,
            "unit_price": 50.0,
            "total": 100.0,
        },
        {
            "brand": "Adidas",
            "model": "Stan Smith",
            "name": "Shoe",
            "barcode": None,
            "qty": 2,
            "unit_price": 50.0,
            "total": 100.0,
        },
    ]
    result = generate_invoice_xlsx_bytes(invoice, items)
    assert result[:2] == b"PK", "Expected XLSX (ZIP) bytes"

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    # Header row is row 5; Barcode is column D (4)
    assert ws.cell(row=5, column=4).value == "Barcode"
    # First data row barcode value
    assert ws.cell(row=6, column=4).value == "1234567890"
    # Second data row barcode is empty (None → empty cell)
    assert not ws.cell(row=7, column=4).value


def test_generate_invoice_xlsx_bytes_missing_barcode_no_crash() -> None:
    """generate_invoice_xlsx_bytes must not crash when items have no barcode key."""
    invoice = {
        "number": 2,
        "client": "Client B",
        "created_at": "2024-02-01T10:00:00",
        "total": 50.0,
    }
    items = [
        {
            "brand": "Puma",
            "model": "RS-X",
            "name": "Sneaker",
            # No "barcode" key at all
            "qty": 1,
            "unit_price": 50.0,
            "total": 50.0,
        },
    ]
    result = generate_invoice_xlsx_bytes(invoice, items)
    assert result[:2] == b"PK", "Expected XLSX (ZIP) bytes"


def test_calc_line_total_uses_displayed_unit_price() -> None:
    assert calc_line_total(8.3033, 3) == pytest.approx(24.90)


def test_generate_invoice_xlsx_bytes_rounds_like_display_price() -> None:
    import io
    import openpyxl

    invoice = {"number": 3, "client": "Client C", "created_at": "2024-03-01T10:00:00", "total": 0}
    items = [
        {
            "brand": "Nike",
            "model": "Air",
            "name": "Sneaker",
            "barcode": "111",
            "qty": 3,
            "unit_price": 8.3033,
            "total": 24.91,
        }
    ]
    wb = openpyxl.load_workbook(io.BytesIO(generate_invoice_xlsx_bytes(invoice, items)))
    ws = wb.active
    assert ws.cell(row=6, column=6).value == pytest.approx(8.30)
    assert ws.cell(row=6, column=7).value == pytest.approx(24.90)
    assert ws.cell(row=7, column=7).value == pytest.approx(24.90)


def test_generate_invoice_xlsx_bytes_sets_print_setup_for_a4_width_fit() -> None:
    import io
    import openpyxl

    invoice = {"number": 4, "client": "Client D", "created_at": "2024-03-01T10:00:00", "total": 0}
    items = [
        {
            "brand": "Nike",
            "model": "Air",
            "name": "Sneaker",
            "barcode": "111",
            "qty": 3,
            "unit_price": 8.3033,
            "total": 24.91,
        }
    ]
    wb = openpyxl.load_workbook(io.BytesIO(generate_invoice_xlsx_bytes(invoice, items)))
    ws = wb.active

    assert str(ws.page_setup.paperSize) == ws.PAPERSIZE_A4
    assert ws.page_setup.orientation == ws.ORIENTATION_PORTRAIT
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True

    assert ws.page_margins.left == pytest.approx(0.25)
    assert ws.page_margins.right == pytest.approx(0.25)
    assert ws.page_margins.top == pytest.approx(0.75)
    assert ws.page_margins.bottom == pytest.approx(0.75)

    # print_area now extends 3 rows past the TOTAL row to include the
    # Seller/Buyer signature line (added in Phase 5). Old expectation was
    # $A$1:$G$7 (header@5 + 1 item + TOTAL@7). Signature sits at TOTAL+3 → G10.
    assert "$A$1:$G$10" in ws.print_area
    assert ws.print_title_rows == "$5:$5"


def test_generate_invoice_xlsx_bytes_uses_left_title_merges_for_stamp_space(
    monkeypatch: pytest.MonkeyPatch, tmp_path
) -> None:
    import io
    import zipfile
    import openpyxl
    from app.services import invoice_xlsx

    monkeypatch.setattr(invoice_xlsx, "STAMP_PATH", tmp_path / "nonexistent.png")

    invoice = {"number": 5, "client": "Client E", "created_at": "2024-03-01T10:00:00", "total": 0}
    items = [
        {
            "brand": "Nike",
            "model": "Air",
            "name": "Sneaker",
            "barcode": "111",
            "qty": 1,
            "unit_price": 10.0,
            "total": 10.0,
        }
    ]
    result = generate_invoice_xlsx_bytes(invoice, items)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    merged = {str(rng) for rng in ws.merged_cells.ranges}

    assert "A1:D1" in merged
    assert "A2:D2" in merged
    assert "A3:D3" in merged
    assert "A1:G1" not in merged
    assert "A2:G2" not in merged
    assert "A3:G3" not in merged
    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        assert not any(name.startswith("xl/media/") for name in zf.namelist())


def test_generate_invoice_xlsx_bytes_inserts_stamp_image_when_png_exists(
    monkeypatch: pytest.MonkeyPatch, tmp_path
) -> None:
    import base64
    import io
    import zipfile
    import xml.etree.ElementTree as ET
    from app.services import invoice_xlsx

    stamp_png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO7+K7sAAAAASUVORK5CYII="
    )
    stamp_path = tmp_path / "stamp.png"
    stamp_path.write_bytes(stamp_png)
    monkeypatch.setattr(invoice_xlsx, "STAMP_PATH", stamp_path)

    invoice = {"number": 6, "client": "Client F", "created_at": "2024-03-01T10:00:00", "total": 0}
    items = [
        {
            "brand": "Nike",
            "model": "Air",
            "name": "Sneaker",
            "barcode": "111",
            "qty": 1,
            "unit_price": 10.0,
            "total": 10.0,
        }
    ]
    result = generate_invoice_xlsx_bytes(invoice, items)
    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        names = zf.namelist()
        media_files = [name for name in names if name.startswith("xl/media/")]
        assert len(media_files) == 1

        drawing_files = [
            name for name in names if name.startswith("xl/drawings/drawing") and name.endswith(".xml")
        ]
        assert drawing_files

        root = ET.fromstring(zf.read(drawing_files[0]))
        ns = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
        from_node = root.find(".//xdr:oneCellAnchor/xdr:from", ns)
        if from_node is None:
            from_node = root.find(".//xdr:twoCellAnchor/xdr:from", ns)

        assert from_node is not None
        assert int(from_node.find("xdr:col", ns).text) == 4  # E
        assert int(from_node.find("xdr:row", ns).text) == 0  # 1


def test_generate_return_xlsx_bytes_rounds_like_display_price() -> None:
    import io
    import openpyxl

    invoice = {
        "number": 7,
        "client": "Client D",
        "created_at": "2024-03-01T10:00:00",
        "warehouse_code": "1416_SHOP",
        "total": 0,
    }
    items = [
        {
            "brand": "Puma",
            "model": "X",
            "name": "Pair",
            "barcode": "222",
            "qty": 3,
            "unit_price": 8.3033,
            "total": 24.91,
        }
    ]
    wb = openpyxl.load_workbook(io.BytesIO(generate_return_xlsx_bytes(invoice, items)))
    ws = wb.active
    assert ws.cell(row=7, column=6).value == pytest.approx(8.30)
    assert ws.cell(row=7, column=7).value == pytest.approx(24.90)
    assert ws.cell(row=8, column=7).value == pytest.approx(24.90)


def test_generate_receive_xlsx_bytes_rounds_like_display_price() -> None:
    import io
    import openpyxl

    invoice = {
        "number": 9,
        "supplier": "Supplier X",
        "created_at": "2024-03-01T10:00:00",
        "destination_warehouse": "TM_DEPO",
        "total": 0,
    }
    items = [
        {
            "brand": "Asics",
            "model": "GEL",
            "name": "Runner",
            "barcode": "333",
            "qty": 3,
            "purchase_price": 8.3033,
            "total": 24.91,
        }
    ]
    wb = openpyxl.load_workbook(io.BytesIO(generate_receive_xlsx_bytes(invoice, items)))
    ws = wb.active
    assert ws.cell(row=7, column=6).value == pytest.approx(8.30)
    assert ws.cell(row=7, column=7).value == pytest.approx(24.90)
    assert ws.cell(row=8, column=7).value == pytest.approx(24.90)


def test_sale_line_item_price_is_rounded_on_save_add_and_update(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    open_cart = db.get_open_cart()
    if open_cart:
        db.cancel_cart(open_cart["cart_id"])

    suffix = uuid.uuid4().hex[:8]
    client_name = f"Round Client {suffix}"
    ok, err = db.add_client(client_name)
    assert ok, err
    client_id = next(c["id"] for c in db.list_clients(include_archived=True) if c["name"] == client_name)
    db.add_warehouse("1416_SHOP", "Shop")

    brand = f"BR{suffix}".upper()
    model = f"m{suffix}".lower()
    product_id = db.add_product(brand, model, "Sale Round Test", 7.0)
    assert product_id > 0

    ok, err, cart_id = db.cart_start_by_id(client_id, "1416_SHOP")
    assert ok, err
    ok, err = db.cart_add_by_cart_id(cart_id, brand, model, 3, "custom", 8.305)
    assert ok, err

    _, items = db.get_cart_items_list(cart_id)
    item = items[0]
    # _normalize_unit_price uses ROUND_HALF_UP (matches Python %.2f and
    # what people expect in trade), so 8.305 rounds up to 8.31.
    assert item["unit_price"] == pytest.approx(8.31)
    assert item["total"] == pytest.approx(24.93)

    ok, err = db.update_cart_item(item["id"], 3, 8.305)
    assert ok, err
    _, items_after = db.get_cart_items_list(cart_id)
    updated = next(i for i in items_after if i["id"] == item["id"])
    assert updated["unit_price"] == pytest.approx(8.31)
    assert updated["total"] == pytest.approx(24.93)

    db.cancel_cart(cart_id)


def test_receive_line_item_price_is_rounded_on_save_add_and_update(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    open_inv = db.receive_invoice_get_open()
    if open_inv:
        db.receive_invoice_cancel(open_inv["id"])

    suffix = uuid.uuid4().hex[:8]
    product_id = db.add_product(f"RB{suffix}".upper(), f"r{suffix}".lower(), "Receive Round Test", 5.0)
    assert product_id > 0
    db.add_warehouse("TM_DEPO", "Depo")

    ok, err, invoice_id = db.receive_invoice_start(f"Supp {suffix}", "TM_DEPO", "")
    assert ok, err

    ok, err = db.receive_item_add(invoice_id, product_id, 3, 8.305)
    assert ok, err

    items = db.receive_invoice_get_items(invoice_id)
    item = items[0]
    # HALF_UP: 8.305 → 8.31
    assert item["purchase_price"] == pytest.approx(8.31)
    assert item["total"] == pytest.approx(24.93)

    ok, err = db.receive_item_update(item["id"], 3, 8.305)
    assert ok, err
    updated = db.receive_invoice_get_items(invoice_id)[0]
    assert updated["purchase_price"] == pytest.approx(8.31)
    assert updated["total"] == pytest.approx(24.93)

    db.receive_invoice_cancel(invoice_id)


def test_return_line_item_price_is_rounded_on_save_add_and_update(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    open_inv = db.return_invoice_get_open()
    if open_inv:
        db.return_invoice_cancel(open_inv["id"])

    suffix = uuid.uuid4().hex[:8]
    client_name = f"Return Round Client {suffix}"
    ok, err = db.add_client(client_name)
    assert ok, err
    client_id = next(c["id"] for c in db.list_clients(include_archived=True) if c["name"] == client_name)
    product_id = db.add_product(f"TB{suffix}".upper(), f"t{suffix}".lower(), "Return Round Test", 5.0)
    assert product_id > 0
    db.add_warehouse("1416_SHOP", "Shop")

    ok, err, invoice_id = db.return_invoice_start(client_id, "1416_SHOP", "")
    assert ok, err
    ok, err = db.return_item_add(invoice_id, product_id, 3, 8.305)
    assert ok, err

    items = db.return_invoice_get_items(invoice_id)
    item = items[0]
    # HALF_UP: 8.305 → 8.31
    assert item["unit_price"] == pytest.approx(8.31)
    assert item["total"] == pytest.approx(24.93)

    ok, err = db.return_item_update(item["id"], 3, 8.305)
    assert ok, err
    updated = db.return_invoice_get_items(invoice_id)[0]
    assert updated["unit_price"] == pytest.approx(8.31)
    assert updated["total"] == pytest.approx(24.93)

    db.return_invoice_cancel(invoice_id)


def test_client_history_renders_datetime_and_debt_signage(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    suffix = uuid.uuid4().hex[:8]
    client_name = f"History Client {suffix}"
    ok, err = db.add_client(client_name)
    assert ok, err
    db.add_warehouse("1416_SHOP", "Shop")
    product_id = db.add_product(f"HB{suffix}".upper(), f"h{suffix}".lower(), "History Product", 10.0)
    assert product_id > 0

    client_id = next(c["id"] for c in db.list_clients(include_archived=True) if c["name"] == client_name)

    with db._connect() as conn:  # noqa: SLF001 - integration setup for deterministic history rows
        next_invoice = conn.execute("SELECT COALESCE(MAX(number), 0) + 1 AS n FROM invoices").fetchone()["n"]
        next_return = conn.execute("SELECT COALESCE(MAX(number), 0) + 1 AS n FROM return_invoices").fetchone()["n"]

        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, status, created_at) VALUES (?, ?, 'CLOSED', ?)",
            (client_id, "1416_SHOP", "2026-01-01 10:00:00"),
        )
        cart_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.execute(
            "INSERT INTO cart_items (cart_id, product_id, qty, price_mode, unit_price, total)"
            " VALUES (?, ?, ?, ?, ?, ?)",
            (cart_id, product_id, 1, "custom", 100.0, 100.0),
        )
        conn.execute(
            "INSERT INTO invoices (cart_id, number, created_at, total) VALUES (?, ?, ?, ?)",
            (cart_id, next_invoice, "2026-01-01 10:00:00", 100.0),
        )
        conn.execute(
            "INSERT INTO return_invoices (number, client_id, warehouse_code, status, created_at, total, note)"
            " VALUES (?, ?, ?, 'DONE', ?, ?, ?)",
            (next_return, client_id, "1416_SHOP", "2026-01-02 10:00:00", 40.0, "Return note"),
        )
        conn.execute(
            "INSERT INTO client_ledger (client_id, created_at, amount, note) VALUES (?, ?, ?, ?)",
            (client_id, "2026-01-03 10:00:00", -331.0, "Debt add"),
        )
        conn.execute(
            "INSERT INTO client_ledger (client_id, created_at, amount, note) VALUES (?, ?, ?, ?)",
            (client_id, "2026-01-04 10:00:00", 165.0, "Debt payment"),
        )
        conn.commit()

    response = client.get(f"/clients/{client_id}/history")
    assert response.status_code == 200
    html = response.text

    assert "2026-01-01 10:00:00" in html
    assert "2026-01-02 10:00:00" in html
    assert "2026-01-03 10:00:00" in html
    assert "2026-01-04 10:00:00" in html

    assert "text-danger\">+331.00" in html
    assert "--331.00" not in html
    assert "text-success\">-165.00" in html
    assert "text-danger\">+100.00" in html
    assert "Balance after" in html
    assert "100.00 USD" in html
    assert "60.00 USD" in html
    assert "391.00 USD" in html


def test_get_client_history_balance_after_is_chronological(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    suffix = uuid.uuid4().hex[:8]
    client_name = f"History Balance {suffix}"
    ok, err = db.add_client(client_name)
    assert ok, err
    db.add_warehouse("1416_SHOP", "Shop")
    product_id = db.add_product(f"BB{suffix}".upper(), f"b{suffix}".lower(), "Balance Product", 10.0)
    assert product_id > 0

    client_id = next(c["id"] for c in db.list_clients(include_archived=True) if c["name"] == client_name)

    with db._connect() as conn:  # noqa: SLF001 - deterministic timeline setup
        next_invoice = conn.execute("SELECT COALESCE(MAX(number), 0) + 1 AS n FROM invoices").fetchone()["n"]

        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, status, created_at) VALUES (?, ?, 'CLOSED', ?)",
            (client_id, "1416_SHOP", "2026-04-16 11:19:42"),
        )
        cart_id_1 = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.execute(
            "INSERT INTO cart_items (cart_id, product_id, qty, price_mode, unit_price, total)"
            " VALUES (?, ?, ?, ?, ?, ?)",
            (cart_id_1, product_id, 1, "custom", 179.13, 179.13),
        )
        conn.execute(
            "INSERT INTO invoices (cart_id, number, created_at, total) VALUES (?, ?, ?, ?)",
            (cart_id_1, next_invoice, "2026-04-16 11:19:42", 179.13),
        )
        conn.execute(
            "INSERT INTO client_ledger (client_id, created_at, amount, note) VALUES (?, ?, ?, ?)",
            (client_id, "2026-04-16 13:35:32", 179.13, "Payment"),
        )

        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, status, created_at) VALUES (?, ?, 'CLOSED', ?)",
            (client_id, "1416_SHOP", "2026-04-18 11:45:36"),
        )
        cart_id_2 = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.execute(
            "INSERT INTO cart_items (cart_id, product_id, qty, price_mode, unit_price, total)"
            " VALUES (?, ?, ?, ?, ?, ?)",
            (cart_id_2, product_id, 1, "custom", 203.92, 203.92),
        )
        conn.execute(
            "INSERT INTO invoices (cart_id, number, created_at, total) VALUES (?, ?, ?, ?)",
            (cart_id_2, next_invoice + 1, "2026-04-18 11:45:36", 203.92),
        )
        conn.execute(
            "INSERT INTO client_ledger (client_id, created_at, amount, note) VALUES (?, ?, ?, ?)",
            (client_id, "2026-04-19 08:58:54", 106.49, "Partial payment"),
        )
        conn.commit()

    events = db.get_client_history(client_id)
    by_dt = {str(ev["created_at"]): float(ev["balance_after"]) for ev in events}
    assert by_dt["2026-04-16 11:19:42"] == pytest.approx(179.13)
    assert by_dt["2026-04-16 13:35:32"] == pytest.approx(0.0)
    assert by_dt["2026-04-18 11:45:36"] == pytest.approx(203.92)
    assert by_dt["2026-04-19 08:58:54"] == pytest.approx(97.43)
    assert float(events[0]["balance_after"]) == pytest.approx(97.43)


def test_supplier_debt_flow_with_receive_and_ledger(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    open_inv = db.receive_invoice_get_open()
    if open_inv:
        db.receive_invoice_cancel(open_inv["id"])

    suffix = uuid.uuid4().hex[:8]
    supplier_name = f"Dealer {suffix}"
    ok, err = db.add_supplier(supplier_name, "+99360000000", "test")
    assert ok, err
    supplier_id = next(s["id"] for s in db.list_suppliers(include_archived=True) if s["name"] == supplier_name)

    db.add_warehouse("TM_DEPO", "Depo")
    product_id = db.add_product(f"SB{suffix}".upper(), f"s{suffix}".lower(), "Supplier Debt Test", 20.0)
    assert product_id > 0

    response = client.post(
        "/receive/start",
        data={
            "supplier_id": supplier_id,
            "destination_warehouse": "TM_DEPO",
            "note": "for supplier debt",
        },
        follow_redirects=False,
    )
    assert response.status_code == 303

    open_inv = db.receive_invoice_get_open()
    assert open_inv is not None
    assert open_inv["supplier_id"] == supplier_id
    assert open_inv["supplier"] == supplier_name

    ok, err = db.receive_item_add(open_inv["id"], product_id, 2, 50.0)
    assert ok, err
    ok, err = db.receive_invoice_finish(open_inv["id"])
    assert ok, err

    assert db.get_supplier_balance(supplier_id) == pytest.approx(100.0)

    ok, err = db.add_supplier_adjustment(supplier_id, 25.0, "payment")
    assert ok, err
    assert db.get_supplier_balance(supplier_id) == pytest.approx(75.0)

    events = db.get_supplier_history(supplier_id)
    assert events
    assert any(ev["kind"] == "RECEIVE" for ev in events)
    assert any(ev["kind"] == "LEDGER" for ev in events)
    assert all("balance_after" in ev for ev in events)

    history_response = client.get(f"/suppliers/{supplier_id}/history")
    assert history_response.status_code == 200
    assert supplier_name in history_response.text
    assert "RECEIVE" in history_response.text


def test_receive_invoice_supplier_name_fallback_without_supplier_id(client: TestClient) -> None:
    from app.db import sqlite as db

    with db._connect() as conn:  # noqa: SLF001 - integration setup
        next_num = conn.execute(
            "SELECT COALESCE(MAX(number), 0) + 1 AS n FROM receive_invoices"
        ).fetchone()["n"]
        conn.execute(
            """
            INSERT INTO receive_invoices
            (number, supplier, supplier_id, destination_warehouse, status, created_at, note, total)
            VALUES (?, ?, NULL, ?, 'DONE', ?, ?, ?)
            """,
            (next_num, "Legacy Supplier", "TM_DEPO", "2026-01-10 10:00:00", "legacy", 42.0),
        )
        conn.commit()

    rows = db.list_receive_invoices_done(q="Legacy Supplier")
    assert rows
    assert rows[0]["supplier"] == "Legacy Supplier"


def test_init_db_backfills_suppliers_from_receive_invoices(client: TestClient) -> None:
    from app.db import sqlite as db

    db.add_warehouse("TM_DEPO", "Depo")
    with db._connect() as conn:  # noqa: SLF001 - integration setup
        next_num = conn.execute(
            "SELECT COALESCE(MAX(number), 0) + 1 AS n FROM receive_invoices"
        ).fetchone()["n"]
        conn.execute(
            """
            INSERT INTO receive_invoices
            (number, supplier, supplier_id, destination_warehouse, status, created_at, note, total)
            VALUES (?, ?, NULL, ?, 'DONE', ?, ?, ?)
            """,
            (next_num, "  Gulzar  ", "TM_DEPO", "2026-01-11 10:00:00", "legacy", 10.0),
        )
        conn.execute(
            """
            INSERT INTO receive_invoices
            (number, supplier, supplier_id, destination_warehouse, status, created_at, note, total)
            VALUES (?, ?, NULL, ?, 'DONE', ?, ?, ?)
            """,
            (next_num + 1, "Yiwu SONIFER", "TM_DEPO", "2026-01-11 11:00:00", "legacy", 20.0),
        )
        conn.commit()

    db.init_db()
    db.init_db()

    with db._connect() as conn:  # noqa: SLF001 - integration assertions
        gulzar_count = conn.execute(
            "SELECT COUNT(*) AS c FROM suppliers WHERE name = 'Gulzar'"
        ).fetchone()["c"]
        yiwu_count = conn.execute(
            "SELECT COUNT(*) AS c FROM suppliers WHERE name = 'Yiwu SONIFER'"
        ).fetchone()["c"]
        assert gulzar_count == 1
        assert yiwu_count == 1

        rows = conn.execute(
            """
            SELECT ri.supplier, ri.supplier_id, s.name AS supplier_name
            FROM receive_invoices ri
            LEFT JOIN suppliers s ON s.id = ri.supplier_id
            WHERE ri.number IN (?, ?)
            ORDER BY ri.number
            """,
            (next_num, next_num + 1),
        ).fetchall()
        assert rows[0]["supplier_id"] is not None
        assert rows[0]["supplier_name"] == "Gulzar"
        assert rows[1]["supplier_id"] is not None
        assert rows[1]["supplier_name"] == "Yiwu SONIFER"


def test_suppliers_nav_label_is_translated(client: TestClient) -> None:
    lang_set = client.post("/set-lang", data={"lang": "ru", "next": "/"}, follow_redirects=False)
    assert lang_set.status_code == 303

    ru_response = client.get("/")
    assert ru_response.status_code == 200
    assert 'href="/suppliers">Поставщики<' in ru_response.text

    lang_set = client.post("/set-lang", data={"lang": "en", "next": "/"}, follow_redirects=False)
    assert lang_set.status_code == 303

    en_response = client.get("/")
    assert en_response.status_code == 200
    assert 'href="/suppliers">Suppliers<' in en_response.text


def test_client_history_empty_state_colspan_matches_visible_columns(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    suffix = uuid.uuid4().hex[:8]
    client_name = f"History Empty {suffix}"
    ok, err = db.add_client(client_name)
    assert ok, err
    client_id = next(c["id"] for c in db.list_clients(include_archived=True) if c["name"] == client_name)

    response = client.get(f"/clients/{client_id}/history")
    assert response.status_code == 200
    html = response.text

    assert "Balance after" in html
    assert 'colspan="7"' in html
    assert 'colspan="6"' not in html
    assert "client-history-table" in html
    assert "col-balance-after" in html
    assert "col-note" in html


def test_invoice_search_filters_all_done_invoice_lists(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    suffix = uuid.uuid4().hex[:8]
    sale_client_name = f"Invoice Search Client {suffix}"
    other_client_name = f"Invoice Other Client {suffix}"
    ok, err = db.add_client(sale_client_name)
    assert ok, err
    ok, err = db.add_client(other_client_name)
    assert ok, err
    db.add_warehouse("1416_SHOP", "Shop")
    db.add_warehouse("TM_DEPO", "Depo")

    clients = {c["name"]: c["id"] for c in db.list_clients(include_archived=True)}
    sale_client_id = clients[sale_client_name]
    other_client_id = clients[other_client_name]

    with db._connect() as conn:  # noqa: SLF001 - integration setup for deterministic invoice rows
        next_sale = conn.execute("SELECT COALESCE(MAX(number), 0) + 1 AS n FROM invoices").fetchone()["n"]
        next_receive = conn.execute(
            "SELECT COALESCE(MAX(number), 0) + 1 AS n FROM receive_invoices"
        ).fetchone()["n"]
        next_return = conn.execute(
            "SELECT COALESCE(MAX(number), 0) + 1 AS n FROM return_invoices"
        ).fetchone()["n"]
        sale_number = next_sale + 1000
        other_sale_number = sale_number + 1
        receive_number = next_receive + 1000
        other_receive_number = receive_number + 1
        return_number = next_return + 1000
        other_return_number = return_number + 1

        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, status, created_at) VALUES (?, ?, 'CLOSED', ?)",
            (sale_client_id, "1416_SHOP", "2026-04-18 10:00:00"),
        )
        sale_cart_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.execute(
            "INSERT INTO invoices (cart_id, number, created_at, total) VALUES (?, ?, ?, ?)",
            (sale_cart_id, sale_number, "2026-04-18 10:00:00", 10.0),
        )
        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, status, created_at) VALUES (?, ?, 'CLOSED', ?)",
            (other_client_id, "1416_SHOP", "2026-04-19 10:00:00"),
        )
        other_cart_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.execute(
            "INSERT INTO invoices (cart_id, number, created_at, total) VALUES (?, ?, ?, ?)",
            (other_cart_id, other_sale_number, "2026-04-19 10:00:00", 11.0),
        )

        conn.execute(
            "INSERT INTO receive_invoices"
            " (number, supplier, destination_warehouse, status, created_at, total, note)"
            " VALUES (?, ?, ?, 'DONE', ?, ?, ?)",
            (receive_number, f"Search Supplier {suffix}", "TM_DEPO", "2026-04-18 11:00:00", 20.0, ""),
        )
        conn.execute(
            "INSERT INTO receive_invoices"
            " (number, supplier, destination_warehouse, status, created_at, total, note)"
            " VALUES (?, ?, ?, 'DONE', ?, ?, ?)",
            (other_receive_number, f"Other Supplier {suffix}", "1416_SHOP", "2026-04-20 11:00:00", 21.0, ""),
        )

        conn.execute(
            "INSERT INTO return_invoices"
            " (number, client_id, warehouse_code, status, created_at, total, note)"
            " VALUES (?, ?, ?, 'DONE', ?, ?, ?)",
            (return_number, sale_client_id, "1416_SHOP", "2026-04-18 12:00:00", 30.0, ""),
        )
        conn.execute(
            "INSERT INTO return_invoices"
            " (number, client_id, warehouse_code, status, created_at, total, note)"
            " VALUES (?, ?, ?, 'DONE', ?, ?, ?)",
            (other_return_number, other_client_id, "TM_DEPO", "2026-04-21 12:00:00", 31.0, ""),
        )
        conn.commit()

    sale_by_date = db.list_sale_invoices_done(q="2026-04-18")
    assert any(inv["number"] == sale_number for inv in sale_by_date)
    assert all(inv["number"] != other_sale_number for inv in sale_by_date)

    sale_by_number = db.list_sale_invoices_done(q=str(sale_number))
    assert any(inv["number"] == sale_number for inv in sale_by_number)
    assert all(inv["number"] != other_sale_number for inv in sale_by_number)

    sale_by_client = db.list_sale_invoices_done(q=sale_client_name)
    assert any(inv["number"] == sale_number for inv in sale_by_client)
    assert all(inv["number"] != other_sale_number for inv in sale_by_client)

    receive_by_supplier = db.list_receive_invoices_done(q=f"Search Supplier {suffix}")
    assert any(inv["number"] == receive_number for inv in receive_by_supplier)
    assert all(inv["number"] != other_receive_number for inv in receive_by_supplier)

    return_by_client = db.list_return_invoices_done(q=sale_client_name)
    assert any(inv["number"] == return_number for inv in return_by_client)
    assert all(inv["number"] != other_return_number for inv in return_by_client)


def test_invoices_page_search_ui_preserves_query_and_tab(client: TestClient) -> None:
    response = client.get("/invoices?tab=sale&q=abc123")
    assert response.status_code == 200
    html = response.text

    assert 'name="q"' in html
    assert 'value="abc123"' in html
    assert 'name="tab" value="sale"' in html
    assert "/invoices?tab=sale&amp;q=abc123" in html
    assert "/invoices?tab=receive&amp;q=abc123" in html
    assert "/invoices?tab=return&amp;q=abc123" in html
    assert '/invoices?tab=sale">Reset</a>' in html


def test_download_allows_invoice_and_backup_files(client: TestClient) -> None:
    from app.services.backup import INVOICES_DIR, BACKUPS_DIR

    INVOICES_DIR.mkdir(parents=True, exist_ok=True)
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)

    invoice_name = "invoice_999999.pdf"
    backup_name = "backup_999999.zip"
    (INVOICES_DIR / invoice_name).write_bytes(b"%PDF-1.4\n")
    (BACKUPS_DIR / backup_name).write_bytes(b"PK\x03\x04")

    invoice_resp = client.get(f"/download?path=invoices/{invoice_name}")
    assert invoice_resp.status_code == 200
    assert invoice_resp.headers.get("content-disposition", "").endswith(f'filename="{invoice_name}"')

    backup_resp = client.get(f"/download?path=backups/{backup_name}")
    assert backup_resp.status_code == 200
    assert backup_resp.headers.get("content-disposition", "").endswith(f'filename="{backup_name}"')


def test_download_rejects_absolute_path_and_traversal(client: TestClient) -> None:
    abs_resp = client.get("/download", params={"path": "/etc/passwd"}, headers={"accept": "application/json"})
    assert abs_resp.status_code == 403
    assert abs_resp.json() == {"error": "download_forbidden"}

    traversal_resp = client.get(
        "/download",
        params={"path": "invoices/../../README.md"},
        headers={"accept": "application/json"},
    )
    assert traversal_resp.status_code == 403
    assert traversal_resp.json() == {"error": "download_forbidden"}

    missing_resp = client.get(
        "/download",
        params={"path": "invoices/does_not_exist.pdf"},
        headers={"accept": "application/json"},
    )
    assert missing_resp.status_code == 404
    assert missing_resp.json() == {"error": "download_not_found"}


def test_admin_backup_create_requires_unlocked_session(client: TestClient) -> None:
    from app.db import sqlite as db
    from app.web.main import _hash_password

    db.set_setting("admin_lock_enabled", "1")
    db.set_setting("site_lock_hash", _hash_password("secret123"))
    db.delete_all_sessions()
    client.cookies.clear()

    locked = client.post("/admin/backup/create", follow_redirects=False)
    assert locked.status_code == 303
    assert locked.headers["location"].startswith("/unlock?next=%2Fadmin%2Fbackup%2Fcreate")

    unlocked = client.post(
        "/unlock",
        data={"password": "secret123", "next": "/admin/backup/create"},
        follow_redirects=False,
    )
    assert unlocked.status_code == 303

    allowed = client.post("/admin/backup/create", follow_redirects=False)
    assert allowed.status_code == 200


def test_get_invoice_by_number_includes_client_id(client: TestClient) -> None:
    """get_invoice_by_number must return client_id so the edit form pre-selects the correct client.

    Previously the SQL query omitted c.client_id from the SELECT, causing
    invoice.client_id to be None and the edit form to default to the first
    client in the list instead of the actual invoice owner.
    """
    import uuid
    from app.db import sqlite as db

    suffix = uuid.uuid4().hex[:8]
    client_name = f"Edit Form Client {suffix}"
    ok, err = db.add_client(client_name)
    assert ok, err
    db.add_warehouse("1416_SHOP", "Shop")
    product_id = db.add_product(f"EB{suffix}".upper(), f"e{suffix}".lower(), "Edit Test Product", 10.0)
    assert product_id > 0

    clients_list = {c["name"]: c["id"] for c in db.list_clients(include_archived=True)}
    the_client_id = clients_list[client_name]

    with db._connect() as conn:  # noqa: SLF001 - integration setup
        next_num = conn.execute("SELECT COALESCE(MAX(number), 0) + 1 AS n FROM invoices").fetchone()["n"]
        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, status) VALUES (?, ?, 'CLOSED')",
            (the_client_id, "1416_SHOP"),
        )
        cart_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.execute(
            "INSERT INTO invoices (cart_id, number, total) VALUES (?, ?, ?)",
            (cart_id, next_num, 0.0),
        )
        conn.commit()

    invoice = db.get_invoice_by_number(next_num)
    assert invoice is not None, "Invoice not found"
    assert "client_id" in invoice, (
        "get_invoice_by_number must include client_id in the returned dict "
        "so that the edit form pre-selects the correct client"
    )
    assert invoice["client_id"] == the_client_id, (
        f"Expected client_id={the_client_id}, got {invoice.get('client_id')}"
    )
    assert invoice["warehouse_code"] == "1416_SHOP"

    # Verify the edit page renders and contains the correct client selected
    response = client.get(f"/invoices/sale/{next_num}/edit")
    assert response.status_code == 200
    html = response.text
    # The option for our client must have `selected` attribute
    assert f'value="{the_client_id}" selected' in html, (
        f"Client {client_name} (id={the_client_id}) must be pre-selected in the edit form"
    )


def test_sale_invoice_edit_accepts_off_stock_item_row(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    suffix = uuid.uuid4().hex[:8]
    warehouse_code = f"ES{suffix[:4]}".upper()
    client_name = f"Edit Sale Off-stock {suffix}"

    ok, err = db.add_client(client_name)
    assert ok, err
    db.add_warehouse(warehouse_code, f"WH {suffix}")
    product_id = db.add_product(f"EB{suffix}".upper(), f"m{suffix}".lower(), "Invoice Edit Product", 10.0)
    assert product_id > 0

    client_id = next(c["id"] for c in db.list_clients(include_archived=True) if c["name"] == client_name)

    with db._connect() as conn:  # noqa: SLF001 - integration setup
        conn.execute(
            "INSERT INTO stock (warehouse_code, product_id, qty) VALUES (?, ?, ?)",
            (warehouse_code, product_id, 10.0),
        )
        next_num = conn.execute("SELECT COALESCE(MAX(number), 0) + 1 AS n FROM invoices").fetchone()["n"]
        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, status) VALUES (?, ?, 'CLOSED')",
            (client_id, warehouse_code),
        )
        cart_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.execute(
            "INSERT INTO invoices (cart_id, number, total) VALUES (?, ?, ?)",
            (cart_id, next_num, 10.0),
        )
        conn.execute(
            "INSERT INTO cart_items (cart_id, product_id, free_line, free_name, qty, price_mode, unit_price, total)"
            " VALUES (?, ?, 0, '', ?, 'custom', ?, ?)",
            (cart_id, product_id, 1.0, 10.0, 10.0),
        )
        conn.commit()

    response = client.post(
        f"/invoices/sale/{next_num}/edit",
        data={
            "client_id": str(client_id),
            "warehouse_code": warehouse_code,
            "product_id": [str(product_id), ""],
            "qty": ["1", "2"],
            "unit_price": ["11.00", "3.50"],
            "free_name": ["", "Delivery service"],
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/invoices?tab=sale&msg=invoice_updated"

    with db._connect() as conn:  # noqa: SLF001 - verification query
        rows = conn.execute(
            "SELECT product_id, free_line, free_name, qty, unit_price"
            " FROM cart_items WHERE cart_id = ? ORDER BY id",
            (cart_id,),
        ).fetchall()
    assert len(rows) == 2
    assert rows[0]["product_id"] == product_id
    assert rows[0]["free_line"] == 0
    assert float(rows[0]["qty"]) == 1.0
    assert float(rows[0]["unit_price"]) == 11.0
    assert rows[1]["product_id"] is None
    assert rows[1]["free_line"] == 1
    assert rows[1]["free_name"] == "Delivery service"
    assert float(rows[1]["qty"]) == 2.0
    assert float(rows[1]["unit_price"]) == 3.5


def test_sale_invoice_edit_rejects_invalid_stock_product_id(client: TestClient) -> None:
    import uuid
    from app.db import sqlite as db

    suffix = uuid.uuid4().hex[:8]
    warehouse_code = f"IV{suffix[:4]}".upper()
    client_name = f"Edit Sale Invalid Product {suffix}"

    ok, err = db.add_client(client_name)
    assert ok, err
    db.add_warehouse(warehouse_code, f"WH {suffix}")
    product_id = db.add_product(f"IB{suffix}".upper(), f"v{suffix}".lower(), "Invoice Edit Invalid Product", 10.0)
    assert product_id > 0

    client_id = next(c["id"] for c in db.list_clients(include_archived=True) if c["name"] == client_name)

    with db._connect() as conn:  # noqa: SLF001 - integration setup
        conn.execute(
            "INSERT INTO stock (warehouse_code, product_id, qty) VALUES (?, ?, ?)",
            (warehouse_code, product_id, 10.0),
        )
        next_num = conn.execute("SELECT COALESCE(MAX(number), 0) + 1 AS n FROM invoices").fetchone()["n"]
        conn.execute(
            "INSERT INTO carts (client_id, warehouse_code, status) VALUES (?, ?, 'CLOSED')",
            (client_id, warehouse_code),
        )
        cart_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.execute(
            "INSERT INTO invoices (cart_id, number, total) VALUES (?, ?, ?)",
            (cart_id, next_num, 10.0),
        )
        conn.execute(
            "INSERT INTO cart_items (cart_id, product_id, free_line, free_name, qty, price_mode, unit_price, total)"
            " VALUES (?, ?, 0, '', ?, 'custom', ?, ?)",
            (cart_id, product_id, 1.0, 10.0, 10.0),
        )
        conn.commit()

    response = client.post(
        f"/invoices/sale/{next_num}/edit",
        data={
            "client_id": str(client_id),
            "warehouse_code": warehouse_code,
            "product_id": ["bad-id"],
            "qty": ["1"],
            "unit_price": ["11.00"],
            "free_name": [""],
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == f"/invoices/sale/{next_num}/edit?msg=invoice_edit_select_product"

    with db._connect() as conn:  # noqa: SLF001 - verification query
        rows = conn.execute(
            "SELECT product_id, free_line, free_name, qty, unit_price"
            " FROM cart_items WHERE cart_id = ? ORDER BY id",
            (cart_id,),
        ).fetchall()
    assert len(rows) == 1
    assert rows[0]["product_id"] == product_id
    assert rows[0]["free_line"] == 0
    assert rows[0]["free_name"] == ""
    assert float(rows[0]["qty"]) == 1.0
    assert float(rows[0]["unit_price"]) == 10.0
