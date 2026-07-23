from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from app.utils.money import calc_document_total, calc_line_total, round_money


OUT_DIR = Path("/opt/stock_bot/invoices")
STAMP_PATH = Path("/opt/stock_bot/stamp.png")  # Optional PNG stamp/logo for sales invoices
STAMP_WIDTH = 220
STAMP_HEIGHT = 80

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="D9EEF7")  # Accent 1, ~60% (light blue)
_TOTAL_FONT = Font(bold=True)
_TOTAL_FONT_ITALIC = Font(bold=True, italic=True)

_CENTER = Alignment(horizontal="center")
_RIGHT = Alignment(horizontal="right")

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _make_workbook(invoice: dict[str, Any], items: list[dict[str, Any]]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Invoice {invoice['number']:06d}"

    # --- Title block ---
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"INVOICE #{invoice['number']:06d}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = _CENTER

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Client: {invoice['client']}"
    ws.merge_cells("A3:D3")
    date_val = str(invoice.get("created_at", invoice.get("date", "")))[:16].replace("T", " ")
    ws["A3"].value = f"Date: {date_val}"
    ws["A4"].value = ""

    # Place optional stamp/logo PNG at /opt/stock_bot/stamp.png (if present).
    if STAMP_PATH.exists():
        stamp = XLImage(str(STAMP_PATH))
        stamp.width = STAMP_WIDTH
        stamp.height = STAMP_HEIGHT
        ws.add_image(stamp, "E1")

    # --- Header row ---
    headers = ["#", "Model", "Name", "Barcode", "Qty", "Unit Price", "Total"]
    col_widths = [5, 18, 30, 18, 8, 14, 14]
    header_row = 5
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[cell.column_letter].width = w

    # --- Data rows ---
    for row_num, item in enumerate(items, start=1):
        row_idx = header_row + row_num
        if item.get("free_line"):
            model_val = item.get("free_name") or item.get("name") or ""
            name_val = ""
        else:
            model_val = f"{item['brand']} {item['model']}"
            name_val = item["name"]
        values = [
            row_num,
            model_val,
            name_val,
            item.get("barcode") or "",
            float(item["qty"]),
            round_money(item["unit_price"]),
            calc_line_total(item["unit_price"], item["qty"]),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            if col_idx in (1, 5):
                cell.alignment = _CENTER
            elif col_idx in (6, 7):
                cell.alignment = _RIGHT

    # --- Total row (styled like your 2nd screenshot) ---
    total_row_idx = header_row + len(items) + 1
    items_qty = int(sum(float(item["qty"]) for item in items))
    invoice_total = calc_document_total(items, "unit_price")

    # Put TOTAL label into Model column (B), bold+italic
    total_lbl = ws.cell(row=total_row_idx, column=2, value="TOTAL")
    total_lbl.font = _TOTAL_FONT_ITALIC
    total_lbl.alignment = Alignment(horizontal="left")
    total_lbl.border = _BORDER

    # Qty sum into Qty column (E)
    qty_cell = ws.cell(row=total_row_idx, column=5, value=items_qty)
    qty_cell.font = _TOTAL_FONT
    qty_cell.alignment = _CENTER
    qty_cell.border = _BORDER

    # Total amount into Total column (G)
    total_cell = ws.cell(row=total_row_idx, column=7, value=invoice_total)
    total_cell.font = _TOTAL_FONT
    total_cell.alignment = _RIGHT
    total_cell.border = _BORDER

    # Keep borders (and light-blue fill) across the whole row A..G
    for col_idx in range(1, 8):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.border = _BORDER
        cell.fill = _TOTAL_FILL

        if col_idx in (1, 5):
            cell.alignment = _CENTER
        elif col_idx in (6, 7):
            cell.alignment = _RIGHT
        else:
    # для текста/пустых ячеек можно left или оставить как есть
    # но лучше задать явно чтобы не было сюрпризов
            cell.alignment = Alignment(horizontal="left")

    # Make sure Unit Price column (F) is empty but styled
    ws.cell(row=total_row_idx, column=6, value="").fill = _TOTAL_FILL
    ws.cell(row=total_row_idx, column=6).border = _BORDER

    # --- Signature block (two side-by-side signatures) ---
    # Leave 2 blank rows, then put Seller on the left half (cols A–C)
    # and Buyer on the right half (cols E–G). Merge each label into a
    # single visual cell so it prints nicely on one line.
    sig_row = total_row_idx + 3
    ws.cell(row=sig_row, column=1, value="Seller: ______________________________").alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=3)
    ws.cell(row=sig_row, column=5, value="Buyer: ______________________________").alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=sig_row, start_column=5, end_row=sig_row, end_column=7)

    # --- Print setup: A4, fit to 1 page wide, multi-page tall ---
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage = True
    ws.page_setup.scale = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Excel "Narrow"-like margins
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    ws.print_area = f"A1:G{sig_row}"
    ws.print_title_rows = f"{header_row}:{header_row}"

    return wb


def generate_invoice_xlsx(invoice: dict[str, Any], items: list[dict[str, Any]]) -> str:
    """Write .xlsx to disk and return the file path."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    number = invoice["number"]
    filepath = OUT_DIR / f"invoice_{number:06d}.xlsx"
    wb = _make_workbook(invoice, items)
    wb.save(str(filepath))
    return str(filepath)


def generate_invoice_xlsx_bytes(invoice: dict[str, Any], items: list[dict[str, Any]]) -> bytes:
    """Return .xlsx content as bytes (for streaming response)."""
    wb = _make_workbook(invoice, items)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
