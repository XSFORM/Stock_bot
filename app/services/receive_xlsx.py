from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="D9EEF7")
_TOTAL_FONT = Font(bold=True)
_TOTAL_FONT_ITALIC = Font(bold=True, italic=True)

_CENTER = Alignment(horizontal="center")
_RIGHT = Alignment(horizontal="right")

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _make_workbook(invoice: dict[str, Any], items: list[dict[str, Any]]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Receive #{invoice['number']:06d}"

    # --- Title block ---
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"PURCHASE INVOICE #{invoice['number']:06d}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = _CENTER

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Supplier: {invoice.get('supplier', '')}"
    ws.merge_cells("A3:G3")
    date_val = str(invoice.get("created_at", ""))[:16].replace("T", " ")
    ws["A3"].value = f"Date: {date_val}"
    ws.merge_cells("A4:G4")
    dest = invoice.get("destination_warehouse", "")
    ws["A4"].value = f"Destination: {dest}"
    ws["A5"].value = ""

    # --- Header row ---
    headers = ["#", "Model", "Name", "Barcode", "Qty", "Purchase Price", "Total"]
    col_widths = [5, 18, 30, 18, 8, 16, 14]
    header_row = 6
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[cell.column_letter].width = w

    # --- Data rows ---
    for row_num, item in enumerate(items, start=1):
        row_idx = header_row + row_num
        values = [
            row_num,
            f"{item['brand']} {item['model']}",
            item["name"],
            item.get("barcode") or "",
            float(item["qty"]),
            round(float(item["purchase_price"]), 2),
            round(float(item["total"]), 2),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            if col_idx in (1, 5):
                cell.alignment = _CENTER
            elif col_idx in (6, 7):
                cell.alignment = _RIGHT

    # --- Total row ---
    total_row_idx = header_row + len(items) + 1
    items_qty = int(sum(float(item["qty"]) for item in items))
    invoice_total = round(sum(float(item["total"]) for item in items), 2)

    total_lbl = ws.cell(row=total_row_idx, column=2, value="TOTAL")
    total_lbl.font = _TOTAL_FONT_ITALIC
    total_lbl.alignment = Alignment(horizontal="left")
    total_lbl.border = _BORDER

    qty_cell = ws.cell(row=total_row_idx, column=5, value=items_qty)
    qty_cell.font = _TOTAL_FONT
    qty_cell.alignment = _CENTER
    qty_cell.border = _BORDER

    total_cell = ws.cell(row=total_row_idx, column=7, value=invoice_total)
    total_cell.font = _TOTAL_FONT
    total_cell.alignment = _RIGHT
    total_cell.border = _BORDER

    for col_idx in range(1, 8):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.border = _BORDER
        cell.fill = _TOTAL_FILL
        if col_idx in (1, 5):
            cell.alignment = _CENTER
        elif col_idx in (6, 7):
            cell.alignment = _RIGHT
        else:
            cell.alignment = Alignment(horizontal="left")

    ws.cell(row=total_row_idx, column=6, value="").fill = _TOTAL_FILL
    ws.cell(row=total_row_idx, column=6).border = _BORDER

    return wb


def generate_receive_xlsx_bytes(invoice: dict[str, Any], items: list[dict[str, Any]]) -> bytes:
    """Return .xlsx content as bytes (for streaming response)."""
    wb = _make_workbook(invoice, items)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
