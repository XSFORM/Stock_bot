from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="D9EEF7")
_TOTAL_FONT = Font(bold=True)

_CENTER = Alignment(horizontal="center")
_RIGHT = Alignment(horizontal="right")

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def generate_stock_xlsx_bytes(rows: list[dict[str, Any]], warehouse_label: str) -> bytes:
    """Generate XLSX bytes for a stock inventory report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    # --- Title ---
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Stock Inventory — {warehouse_label}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = _CENTER

    # --- Header row ---
    headers = ["Brand", "Model", "Name", "Warehouse", "Qty"]
    col_widths = [16, 16, 30, 14, 8]
    header_row = 3
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[cell.column_letter].width = w

    # --- Data rows ---
    for row_num, r in enumerate(rows, start=1):
        row_idx = header_row + row_num
        values = [
            r["brand"],
            r["model"],
            r["name"],
            r.get("warehouse") if "warehouse" in r else r.get("warehouse_code", warehouse_label),
            int(r["qty"]),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            if col_idx == 5:
                cell.alignment = _RIGHT

    # --- Total row ---
    total_row_idx = header_row + len(rows) + 1
    total_qty = sum(int(r["qty"]) for r in rows)

    for col_idx in range(1, 6):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.border = _BORDER

    lbl = ws.cell(row=total_row_idx, column=1, value="TOTAL")
    lbl.font = _TOTAL_FONT

    qty_cell = ws.cell(row=total_row_idx, column=5, value=total_qty)
    qty_cell.font = _TOTAL_FONT
    qty_cell.alignment = _RIGHT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
